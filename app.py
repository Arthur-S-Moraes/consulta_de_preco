from datetime import datetime
from time import sleep
import schedule
import openpyxl
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as condicao_esperada
from selenium.common.exceptions import *

def projeto_monitorar_preco():
    def iniciar_driver():
        print('iniciando o programa, aguarde alguns segundos')
        chrome_option = Options()
        arguments = ['--lang=en-US', '--window-size=1080,1080', '--force-dark-mode', '--log-level=3']
        for argument in arguments:
            chrome_option.add_argument(argument)


        driver = webdriver.Chrome(service=ChromeService(
            ChromeDriverManager().install()), options=chrome_option)
        
        wait = WebDriverWait(
            driver,
            10,
            poll_frequency=1,
            ignored_exceptions=[
                NoSuchElementException,
                ElementNotVisibleException,
                ElementNotSelectableException
            ]
        )
        return driver, wait
    def buscar_produto(site='https://www.kabum.com.br/produto/609733/console-playstation-5-slim-digital-edition-branco-returnal-e-ratchet-e-clank-controle-sem-fio-dualsense-branco'):
        site
        driver, wait = iniciar_driver()
        print('iniciando o navegador')
        try:
            driver.get(site)
            print('buscando o valor do produto selecionado')
            produto = wait.until(condicao_esperada.visibility_of_element_located((By.XPATH, "//h1[@class='sc-58b2114e-6 brTtKt']"))).text
            valor = wait.until(condicao_esperada.visibility_of_element_located((By.XPATH, "//h4[@class='sc-5492faee-2 ipHrwP finalPrice']"))).text
            valor = valor.split(' ')[1]
        except:
            print('''tivemos algum erro e não foi possivel encontrar o valor do item selecionado
                use uma URL com o produto que deseja buscar o valor''')
        return(produto, valor, site)
    def criar_editar_planilha():
        produto, valor, site = buscar_produto()

        try:
            workbook = openpyxl.load_workbook('produtos.xlsx')
            sheet_produtos = workbook['Produtos']
            print('carregando e atualizando a planilha salva')
        except:
            workbook = openpyxl.Workbook()
            print('criando nova planilha')
            workbook['Sheet'].title = 'Produtos'
            sheet_produtos = workbook['Produtos']
            sheet_produtos.append(['Produto', 'Data atual', 'Valor', 'Link'])
        data_atual = datetime.now()
        data_formatada = data_atual.strftime('%d/%m/%Y %H:%M')

        for linha in sheet_produtos.iter_rows(min_row=2, max_row=100,min_col=1, max_col=4):
            if linha[0].value == produto and linha[3].value == site:
                linha[1].value = data_formatada
                linha[2].value = valor
                break
            elif linha[0].value == None:
                linha[0].value = produto
                linha[1].value = data_formatada
                linha[2].value = valor
                linha[3].value = site
                break

        workbook.save('produtos.xlsx')
    criar_editar_planilha()


schedule.every(30).minutes.do(projeto_monitorar_preco)

print('Bot agendado para rodar a cada 30 minutos. pressione Ctrl+C para interromper.')

while True:
    schedule.run_pending()
    sleep(1)